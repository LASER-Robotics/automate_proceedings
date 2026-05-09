"""
Input: camera_ready_report.csv from camera_ready_check.py and SearchCopyright.xlsx from IEEE eCF Management Toolkit

Usage:
    python3 ecf_compliance_check.py --extracted camera_ready_report.csv --copyright SearchCopyright.xlsx
"""
import argparse
import difflib
import re
import unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import warnings
import csv

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- Visual Styling Configuration ---
COLORS = {
    "HEADER_BG": "1F3864", "HEADER_FT": "FFFFFF",
    "OK_BG": "E2EFDA", "WARNING_BG": "FFF2CC", "ERROR_BG": "FCE4D6",
    "BORDER": "CCCCCC", "SUCCESS_TEXT": "375623", "ALERT_TEXT": "843C0C"
}

# ---------------------------------------------------------------------------
# Text Normalization & Noise Filtering
# ---------------------------------------------------------------------------

def clean_noise(text: str) -> str:
    """Removes academic titles, technical terms, and common address noise."""
    if pd.isna(text): return ""
    t = str(text).lower()
    
    # Common artifacts leaked from PDF metadata/headers
    patterns = [
        r"\bav\b", r"\bavenida\b", r"\brua\b", r"\bcampus\b", r"\bbloco\b",
        r"\buniversidade\b", r"\bufpb\b", r"\bufrj\b", r"\busp\b", r"\bcefet\b",
        r"\bprof\b", r"\bdr\b", r"\bphd\b", r"\bprofessor\b", 
        r"\bhowever\b", r"\bintroduction\b", r"\babstract\b", 
        r"\bartificial intelligence\b", r"\bzenite\b", r"\bi\.\s*i\b"
    ]
    
    for p in patterns:
        t = re.sub(p, "", t)
    return t

def fix_pdf_encoding(text: str) -> str:
    """Corrects broken encoding sequences common in PDF text extraction."""
    t = text.lower()
    mapping = [
        ("a˜", "ao"), ("o˜", "ao"), ("c¸", "c"), ("a´", "a"), ("e´", "e"),
        ("i´", "i"), ("o´", "o"), ("u´", "u"), ("´i", "i"), ("´a", "a"),
        ("´e", "e"), ("´o", "o"), ("´u", "u"), ("˜a", "a"), ("˜o", "o")
    ]
    for broken, fixed in mapping:
        t = t.replace(broken, fixed)
    return t

def normalize_text(text: str, smashed: bool = False) -> str:
    """Standardizes text for robust comparison by removing accents and symbols."""
    t = fix_pdf_encoding(clean_noise(text))
    # Decompose unicode characters and strip non-alphabetic symbols
    n = unicodedata.normalize("NFKD", t).encode("ASCII", "ignore").decode("utf-8")
    n = re.sub(r"[^a-z\s]", "", n)
    
    if smashed:
        return n.replace(" ", "")
    return " ".join(n.split()).strip()

# ---------------------------------------------------------------------------
# Core Matching Logic
# ---------------------------------------------------------------------------

def is_same_person(pdf_name: str, form_name: str) -> bool:
    """Matches names using smashed-containment and initial-based heuristics."""
    s_pdf = normalize_text(pdf_name, smashed=True)
    s_form = normalize_text(form_name, smashed=True)
    
    if not s_pdf or not s_form: return False

    # Check for exact smashed match or string containment (handles middle names/suffixes)
    if s_pdf == s_form: return True
    if len(s_pdf) > 5 and len(s_form) > 5:
        if s_pdf in s_form or s_form in s_pdf:
            return True

    # Check for initial-based matching
    n_pdf, n_form = normalize_text(pdf_name), normalize_text(form_name)
    w_pdf, w_form = n_pdf.split(), n_form.split()
    
    if len(w_pdf) < 2 or len(w_form) < 2: return False

    def check_initials(short, long):
        used = set()
        for sw in short:
            found = False
            for idx, lw in enumerate(long):
                if idx not in used and (lw.startswith(sw) or sw.startswith(lw)):
                    used.add(idx)
                    found = True
                    break
            if not found: return False
        return True

    return check_initials(w_pdf, w_form) or check_initials(w_form, w_pdf)

# ---------------------------------------------------------------------------
# Author Diagnosis & Data Processing
# ---------------------------------------------------------------------------

def run_diagnosis(pdf_raw: str, form_raw: str) -> tuple:
    """Compares author lists and identifies missing or extra contributors."""
    p_names = [n.strip() for n in str(pdf_raw).split(",") if len(normalize_text(n)) > 2]
    f_names = [n.strip() for n in str(form_raw).split(",") if len(normalize_text(n)) > 2]

    if not p_names or "[error" in str(pdf_raw).lower():
        return 0.0, "INCONCLUSIVE: Extraction failed in PDF"

    p_idx = list(range(len(p_names)))
    f_idx = list(range(len(f_names)))

    # Step 1: Individual Smart Match
    for p in list(p_idx):
        for f in list(f_idx):
            if is_same_person(p_names[p], f_names[f]):
                p_idx.remove(p)
                f_idx.remove(f)
                break

    # Step 2: Name Fusion (merges split name fragments like 'First Middle' + 'Last')
    for _ in range(2):
        to_del_p, to_del_f = set(), set()
        for i in range(len(p_idx)):
            for span in [2, 3]:
                if i + span > len(p_idx): continue
                fused = " ".join([p_names[idx] for idx in p_idx[i : i+span]])
                for f in f_idx:
                    if f in to_del_f: continue
                    if is_same_person(fused, f_names[f]):
                        to_del_p.update(p_idx[i : i+span])
                        to_del_f.add(f)
                        break
        p_idx = [idx for idx in p_idx if idx not in to_del_p]
        f_idx = [idx for idx in f_idx if idx not in to_del_f]

    score = round(((len(f_names) - len(f_idx)) / len(f_names)) * 100, 2) if f_names else 100.0
    
    logs = []
    for p in p_idx: logs.append(f"ADDED (PDF): {p_names[p]}")
    for f in f_idx: logs.append(f"REMOVED (Form): {f_names[f]}")
    
    return score, ("OK" if not logs else " | ".join(logs))

# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

def save_report(data, path):
    """Generates a professionally styled Excel report with conditional formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"
    
    headers = [
        "ID", "CMT ID", "ARTICLE ID", "ACTION REQUIRED?", "PDF TITLE", "FORM TITLE", 
        "TITLE SCORE (%)", "PDF AUTHORS", "FORM AUTHORS", "AUTHOR SCORE (%)", "DIAGNOSIS"
    ]
    
    side = Side(style="thin", color=COLORS["BORDER"])
    border = Border(left=side, right=side, top=side, bottom=side)

    # Styling Headers
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(name="Arial", bold=True, color=COLORS["HEADER_FT"])
        cell.fill = PatternFill("solid", start_color=COLORS["HEADER_BG"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Styling Rows
    for row_i, item in enumerate(data, 2):
        row_vals = [item[k] for k in ["p_id", "file", "id", "action", "t_pdf", "t_form", "s_title", "a_pdf", "a_form", "s_auth", "diag"]]
        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            
            if col_i == 4: # Action Column
                bg = COLORS["OK_BG"] if val == "NO" else (COLORS["WARNING_BG"] if val == "INCONCLUSIVE" else COLORS["ERROR_BG"])
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(name="Arial", bold=True, color=(COLORS["SUCCESS_TEXT"] if val == "NO" else COLORS["ALERT_TEXT"]))
            elif col_i in [7, 10]: # Score Columns
                score_bg = COLORS["OK_BG"] if val >= 95 else (COLORS["WARNING_BG"] if val >= 80 else COLORS["ERROR_BG"])
                cell.fill = PatternFill("solid", start_color=score_bg)

    # Column Widths
    widths = [12, 15, 18, 50, 50, 15, 50, 50, 15, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = "A2"
    wb.save(path.replace("csv","xlsx"))

    sheet = wb.active
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

def signed_copyright(t_pdf, copyright_xlsx):
    df_cpy = pd.read_excel(copyright_xlsx)
    # Match Article Titles
    t_pdf_sm = normalize_text(t_pdf, smashed=True)
    match_row, best_score = None, -1
    for _, c_row in df_cpy.iterrows():
        c_title_sm = normalize_text(str(c_row.get("ARTICLE TITLE", "")), smashed=True)
        score = difflib.SequenceMatcher(None, t_pdf_sm, c_title_sm).ratio()
        if score > best_score:
            best_score, match_row = score, c_row
    # Returns Y or N
    return match_row.get("COPYRIGHT TYPE", "")

# ---------------------------------------------------------------------------
# Main Orchestrator
# ---------------------------------------------------------------------------

def run_pipeline(extracted_xlsx, copyright_xlsx, output_xlsx):
    df_ext = pd.read_csv(extracted_xlsx)
    df_cpy = pd.read_excel(copyright_xlsx)

    count = 1
    
    results = []
    for _, row in df_ext.iterrows():
        proceedings_id = f"{count:03d}.pdf"
        count = count + 1
        t_pdf, a_pdf = str(row.get("title_pdf", "")), str(row.get("authors_pdf", ""))
        if not t_pdf.strip(): continue

        # Match Article Titles
        t_pdf_sm = normalize_text(t_pdf, smashed=True)
        match_row, best_score = None, -1
        for _, c_row in df_cpy.iterrows():
            c_title_sm = normalize_text(str(c_row.get("ARTICLE TITLE", "")), smashed=True)
            score = difflib.SequenceMatcher(None, t_pdf_sm, c_title_sm).ratio()
            if score > best_score:
                best_score, match_row = score, c_row

        if match_row is not None:
            s_auth, diag = run_diagnosis(a_pdf, match_row["AUTHORS"])
            s_title = round(best_score * 100, 2)
            
            if "INCONCLUSIVE" in diag: action = "INCONCLUSIVE"
            else: action = "NO" if (diag == "OK" and s_title >= 75) else "YES"
            
            results.append({
                "p_id": proceedings_id, "file": row["cmt_id"], "id": match_row.get("ARTICLE IDENTIFIER", ""),
                "action": action, "t_pdf": t_pdf, "t_form": match_row["ARTICLE TITLE"],
                "s_title": s_title, "a_pdf": a_pdf, "a_form": match_row["AUTHORS"],
                "s_auth": s_auth, "diag": diag
            })

    save_report(results, output_xlsx)
    print(f"Compliance analysis completed successfully: {output_xlsx}")

if __name__ == "__main__":
    output_folder = "./reports/"
    input_folder = "./input_data/"
    parser = argparse.ArgumentParser(description="Copyright Compliance Validation Tool")
    parser.add_argument("--extracted", default="extracted_articles.csv", help="Path to extracted PDF data")
    parser.add_argument("--copyright", default="SearchCopyright.xlsx", help="Path to official copyright DB")
    parser.add_argument("--output", default="final_compliance_report.csv", help="Output filename")
    args = parser.parse_args()
    
    run_pipeline(output_folder + args.extracted, input_folder + args.copyright, output_folder + args.output)