"""
RAS Format Extractor & Validator
=================================
Scans a folder of PDF files and, for each one:
  - Verifies structural compliance with the IEEE/RAS ieeeconf template
    (http://ras.papercept.net/conferences/support/files/ieeeconf.zip)
  - Extracts the article TITLE and AUTHORS when the format is valid
  - Records detailed non-compliance reasons when it is not

Key rule for the author block (from root.tex / ieeeconf standard):
  CORRECT   -> author{Name1 and Name2 thanks{...affiliations as footnotes...}}
               Authors appear on ONE centered line; affiliations go to page footer.
  INCORRECT -> IEEEauthorblockN / IEEEauthorblockA
               Produces multi-column blocks with dept/city/email in the header area.

Detection heuristic: if affiliation-like text appears within 60pt BELOW the
last author line, the paper used the wrong author-block format.

Font sizes in the compiled PDFs of these articles: title ~23.9 pt, authors 11 pt,
body 10 pt, abstract labels 9 pt. The ieeeconf.cls profiles below are kept for
reference; actual detection uses tolerant matching.

Usage
-----
    python ras_format_validation.py --folder "path/to/pdfs"
    python ras_format_validation.py --folder "path/to/pdfs" --output results.xlsx
"""

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Visual Styling
# ---------------------------------------------------------------------------

COLORS = {
    "HEADER_BG": "1F3864", "HEADER_FT": "FFFFFF",
    "OK_BG":     "E2EFDA", "OK_FT":     "375623",
    "WARN_BG":   "FFF2CC", "WARN_FT":   "7B6000",
    "ERROR_BG":  "FCE4D6", "ERROR_FT":  "843C0C",
    "BORDER":    "CCCCCC",
}

# ---------------------------------------------------------------------------
# ieeeconf.cls font-size profiles (for reference / base-size detection)
# ---------------------------------------------------------------------------
_CLS_PROFILES: dict[int, dict[str, float]] = {
    9:  {"LARGE": 14.0, "sublargesize": 10.0, "normalsize":  9.0, "small":  8.5},
    10: {"LARGE": 16.0, "sublargesize": 11.0, "normalsize": 10.0, "small":  9.0},
    11: {"LARGE": 17.0, "sublargesize": 12.0, "normalsize": 11.0, "small": 10.0},
    12: {"LARGE": 20.0, "sublargesize": 14.0, "normalsize": 12.0, "small": 10.0},
}

# Tolerance for font-size comparison (PDF renderers vary slightly)
_FONT_TOL = 0.8

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------
PAGE_W_PT     = 612.0
PAGE_H_PT     = 792.0
PAGE_TOL_PT   =   5.0
MARGIN_MIN_PT =  40.0

# Author-block detection
AUTHOR_SIZE_MIN  = 10.0   # author names are >= 10 pt
AUTHOR_SIZE_MAX  = 13.0
TITLE_SIZE_MIN   = 18.0   # titles are clearly large
AFFIL_SCAN_PT    = 65.0   # pt below last author line to look for affiliations
BODY_SIZE        = 10.0   # normal body text size (kept for reference)
BODY_TOL         = 0.6

# Affiliation font range: IEEEauthorblockA renders affiliations between
# ~7pt and ~12pt depending on the compiler and base font size.
# This wider range replaces the old BODY_SIZE +/- BODY_TOL filter.
AFFIL_SIZE_MIN   = 7.0
AFFIL_SIZE_MAX   = 12.0

ABSTRACT_SEARCH_CHARS = 3000
TITLE_SIZE_MIN   = 13.0   # title must be at least this large (absolute floor)

# ---------------------------------------------------------------------------
# Affiliation keywords (used to detect \IEEEauthorblockA misuse)
# ---------------------------------------------------------------------------
_AFFIL_KEYWORDS = {
    "university", "universidade", "universidad", "université",
    "institute", "instituto", "department", "departamento",
    "federal", "nacional", "national", "laboratory", "laboratório",
    "center", "centre", "centro", "school", "escola", "faculty",
    "faculdade", "college", "campus", "engineering", "engenharia",
    "science", "ciência", "technology", "research", "pesquisa",
    "programa", "graduate", "dept",
    # cities/countries that appear in affiliation blocks
    "brazil", "brasil", "germany", "france", "portugal", "spain",
    "italia", "japan", "china", "usa", "netherlands",
}
_PREP = {"da", "de", "do", "das", "dos", "e", "van", "von", "di", "del"}

RE_EMAIL = re.compile(r"[\w.\-+]+@[\w.\-]+\.\w+", re.IGNORECASE)
RE_ORCID = re.compile(r"\d{4}-\d{4}-\d{4}-\d{3}[\dX]")
RE_SUPER = re.compile(r"([A-Za-zÀ-ÿ])[\d*†‡§∗¶#,]+")

# ---------------------------------------------------------------------------
# Encoding repair
# ---------------------------------------------------------------------------
_DOTLESS_I = "\u0131"
_MODS = set("´˜ˆ¸")
_ENC_ANTES = {
    ("a","´"):"á",("e","´"):"é",("o","´"):"ó",("u","´"):"ú",
    ("A","´"):"Á",("E","´"):"É",("O","´"):"Ó",("U","´"):"Ú",
    ("a","˜"):"ã",("o","˜"):"õ",("n","˜"):"ñ",
    ("A","˜"):"Ã",("O","˜"):"Õ",("N","˜"):"Ñ",
    ("e","ˆ"):"ê",("a","ˆ"):"â",("o","ˆ"):"ô",
    ("E","ˆ"):"Ê",("A","ˆ"):"Â",("O","ˆ"):"Ô",
    ("c","¸"):"ç",("C","¸"):"Ç",
}
_ENC_DEPOIS = {
    "´":{"a":"á","e":"é","i":"í","o":"ó","u":"ú","A":"Á","E":"É","I":"Í","O":"Ó","U":"Ú"},
    "˜":{"a":"ã","o":"õ","n":"ñ","A":"Ã","O":"Õ"},
    "ˆ":{"a":"â","e":"ê","o":"ô","A":"Â","E":"Ê","O":"Ô"},
    "¸":{"c":"ç","C":"Ç"},
}

def _fix_encoding(text: str) -> str:
    t = text.replace("´" + _DOTLESS_I, "í")
    chars, out, i = list(t), [], 0
    while i < len(chars):
        ch  = chars[i]
        nxt = chars[i+1] if i+1 < len(chars) else ""
        if nxt in _MODS and (ch, nxt) in _ENC_ANTES:
            out.append(_ENC_ANTES[(ch, nxt)]); i += 2
        else:
            out.append(ch); i += 1
    t = "".join(out)
    for mod, mapping in _ENC_DEPOIS.items():
        t = re.sub(re.escape(mod) + r"(.)",
                   lambda m, mp=mapping: mp.get(m.group(1), m.group(1)), t)
    return re.sub(r"[´˜ˆ¸`¨\u0131]", "", t)

# ---------------------------------------------------------------------------
# Data structure
# ---------------------------------------------------------------------------
@dataclass
class ValidationResult:
    file:    str
    status:  str = "VALID"
    title:   str = ""
    authors: str = ""
    issues:  list = field(default_factory=list)

    @property
    def issues_text(self):
        return " | ".join(self.issues) if self.issues else "OK"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _sz(w) -> float:
    return w.get("size", 0) or 0.0

def _matches(size: float, target: float, tol: float = _FONT_TOL) -> bool:
    return abs(size - target) <= tol

def _is_affil_word(word: str) -> bool:
    """Return True if this single word looks like part of an affiliation."""
    if RE_EMAIL.search(word) or RE_ORCID.search(word):
        return True
    clean = re.sub(r"[^a-záàâãéèêíïóôõúüçñ]", "", word.lower())
    return clean in _AFFIL_KEYWORDS and clean not in _PREP

def _line_is_affil(line_text: str) -> bool:
    """Return True if the line contains affiliation keywords."""
    if RE_EMAIL.search(line_text) or RE_ORCID.search(line_text):
        return True
    words = re.split(r"[\s,.()\-/]+", line_text)
    for w in words:
        if _is_affil_word(w):
            return True
    return False

def _group_lines(words: list, tol: float = 3.0) -> list[list]:
    if not words:
        return []
    lines, cur = [], [words[0]]
    for w in words[1:]:
        if abs(w["top"] - cur[-1]["top"]) <= tol:
            cur.append(w)
        else:
            lines.append(cur)
            cur = [w]
    lines.append(cur)
    return lines

def _detect_base_size(words: list) -> tuple:
    freq = Counter(round(_sz(w), 1) for w in words if _sz(w) > 0)
    for size, _ in freq.most_common():
        for base, profile in _CLS_PROFILES.items():
            if _matches(size, profile["normalsize"]):
                return base, size
    return None, None

# ---------------------------------------------------------------------------
# CHECK 1 — Page size
# ---------------------------------------------------------------------------
def _check_page_size(page, issues: list) -> bool:
    w_ok = abs(page.width  - PAGE_W_PT) <= PAGE_TOL_PT
    h_ok = abs(page.height - PAGE_H_PT) <= PAGE_TOL_PT
    if not w_ok or not h_ok:
        issues.append(
            f"Tamanho de página incorreto: {page.width:.0f}×{page.height:.0f} pt "
            f"(exigido US Letter {PAGE_W_PT:.0f}×{PAGE_H_PT:.0f} pt). "
            "ieeeconf suporta apenas letterpaper — A4 produz margens diferentes."
        )
    return w_ok and h_ok

# ---------------------------------------------------------------------------
# CHECK 2 — Two-column layout
# ---------------------------------------------------------------------------
def _check_two_column(words: list, page_width: float, issues: list) -> bool:
    mid   = page_width / 2
    left  = sum(1 for w in words if w["x1"] < mid + 20)
    right = sum(1 for w in words if w["x0"] > mid - 20)
    ok    = left > 10 and right > 10
    if not ok:
        issues.append(
            f"Layout não parece duas colunas (esquerda: {left} palavras, direita: {right} palavras). "
            "ieeeconf exige modo twocolumn com columnsep de 0,2 in."
        )
    return ok

# ---------------------------------------------------------------------------
# CHECK 3 — Side margins
# ---------------------------------------------------------------------------
def _check_margins(words: list, page_width: float, issues: list) -> bool:
    if not words:
        issues.append("Nenhum texto encontrado — não foi possível verificar margens.")
        return False
    lm = min(w["x0"] for w in words)
    rm = page_width - max(w["x1"] for w in words)
    ok = lm >= MARGIN_MIN_PT and rm >= MARGIN_MIN_PT
    if not ok:
        issues.append(
            f"Margens muito estreitas — esquerda: {lm:.1f} pt, direita: {rm:.1f} pt. "
            f"ieeeconf especifica 0,75 in (54 pt) por lado; mínimo aceito: {MARGIN_MIN_PT} pt."
        )
    return ok

# ---------------------------------------------------------------------------
# CHECK 4 — Abstract keyword
# ---------------------------------------------------------------------------
def _check_abstract(text: str, issues: list) -> bool:
    ok = bool(re.search(r"\bAbstract\b", text[:ABSTRACT_SEARCH_CHARS], re.IGNORECASE))
    if not ok:
        issues.append(
            "Palavra-chave 'Abstract' não encontrada na página 1. "
            "ieeeconf exige \\begin{abstract}...\\end{abstract} na primeira página."
        )
    return ok

# ---------------------------------------------------------------------------
# CHECK 5 — Title font (large, spanning full width)
# ---------------------------------------------------------------------------
def _check_title_font(words: list, page_height: float,
                      page_width: float, issues: list) -> tuple[bool, float | None]:
    """
    The title uses the LARGEST font on the page (ieeeconf: \\LARGE \\bf).
    It must appear in the upper half and span >= 50% of page width
    (confirming it crosses both columns, not a column-level heading).

    We detect title_size as the maximum font size present in the upper half,
    provided it is at least TITLE_SIZE_MIN pt.
    """
    upper_words = [w for w in words if w.get("top", 0) < page_height * 0.50 and _sz(w) > 0]
    if not upper_words:
        issues.append("Nenhum texto encontrado na metade superior da página 1.")
        return False, None

    max_size = max(_sz(w) for w in upper_words)
    if max_size < TITLE_SIZE_MIN:
        issues.append(
            f"Maior fonte na metade superior da página: {max_size:.1f} pt "
            f"(mínimo esperado para título: {TITLE_SIZE_MIN} pt). "
            "Verifique \\title{\\LARGE \\bf ...} no .tex."
        )
        return False, None

    # Collect all words at the title size (with tolerance)
    tw = [w for w in upper_words if _matches(_sz(w), max_size, tol=0.5)]

    title_size = round(max_size, 1)
    span  = max(w["x1"] for w in tw) - min(w["x0"] for w in tw)
    ratio = span / page_width

    if ratio < 0.50:
        issues.append(
            f"Título ocupa apenas {ratio*100:.0f}% da largura da página (mínimo 50%). "
            "No ieeeconf o bloco de título se estende pelas duas colunas."
        )
        return False, title_size

    return True, title_size

# ---------------------------------------------------------------------------
# CHECK 6 — Author line exists (11 pt, single centered line)
# ---------------------------------------------------------------------------
def _find_author_line(words: list, title_size: float,
                      page_height: float) -> tuple[list, float | None]:
    """
    Find the author-name words: they sit just below the title, use a font
    between AUTHOR_SIZE_MIN and AUTHOR_SIZE_MAX, and appear before the abstract
    (top 35% of page).
    """
    title_bottom = max(
        (w["bottom"] for w in words if _matches(_sz(w), title_size, tol=1.5)),
        default=0,
    )
    h35 = page_height * 0.35
    aw = [w for w in words
          if w.get("top", 0) > title_bottom
          and w.get("top", 0) < h35
          and AUTHOR_SIZE_MIN <= _sz(w) <= AUTHOR_SIZE_MAX]

    if not aw:
        return [], None

    last_top = max(w["top"] for w in aw)
    return aw, last_top

# ---------------------------------------------------------------------------
# CHECK 7 — Author block format (thanks vs IEEEauthorblock)
# ---------------------------------------------------------------------------
def _check_author_format(words: list, last_author_top: float,
                         page_height: float, issues: list) -> bool:
    """
    In the CORRECT format (\\author{...\\thanks{...}}), the area just below
    the author line contains body text (10 pt running prose) or nothing —
    affiliations go to the page footer via \\thanks{}.

    In the WRONG format (IEEEauthorblockN/A), the area just below the
    author line contains affiliation text (dept, university, city, email)
    still in the header region. These affiliations are typically rendered
    between 7pt and 12pt depending on the compiler and base font size.

    We scan AFFIL_SCAN_PT below the last author line for words in that size
    range; if those words contain affiliation keywords the format is wrong.

    Special case: some IEEEauthorblockN/A layouts render affiliations at the
    same font size as author names (11pt), making them indistinguishable by
    size alone. In that case the affiliation keywords appear directly inside
    the author_words block itself, so we also check that block.
    """
    _AFFIL_MSG = (
        "Formato de autores incorreto: afiliações (departamento/universidade/cidade/e-mail) "
        "foram colocadas no cabeçalho junto com os nomes, usando \\IEEEauthorblockN/A. "
        "O padrão RAS exige \\author{Nomes \\thanks{afiliações como notas de rodapé}}. "
        "As afiliações devem aparecer no rodapé da primeira página, não no cabeçalho."
    )

    # CHANGED (1): also inspect the author_words block itself for affiliation
    # keywords — handles cases where IEEEauthorblockA renders at 11pt (same
    # size as author names), so affiliations are captured inside author_words
    # and never appear "below" the last_author_top window.
    author_block = [w for w in words if AFFIL_SIZE_MIN <= _sz(w) <= AFFIL_SIZE_MAX
                    and w.get("top", 0) <= last_author_top]
    author_block_sorted = sorted(author_block, key=lambda w: (round(w["top"]), w["x0"]))
    for line in _group_lines(author_block_sorted, tol=4.0):
        line_text = " ".join(w["text"] for w in line)
        if _line_is_affil(line_text):
            issues.append(_AFFIL_MSG)
            return False

    scan_end = last_author_top + AFFIL_SCAN_PT

    # Determine page mid-point to exclude right-column body text.
    # In the correct \thanks{} format, two-column body prose starts in the
    # right column immediately below the author line; IEEEauthorblockA
    # affiliations are centered and therefore span both columns (or only
    # the left for single-author papers). Restricting the scan to the left
    # half prevents right-column body words (e.g. "League", "University"
    # in running prose) from triggering false affiliation matches.
    page_mid = max((w["x1"] for w in words), default=612.0) / 2

    # CHANGED (2): replaced abs(_sz(w) - BODY_SIZE) <= BODY_TOL with a wider
    # range (AFFIL_SIZE_MIN..AFFIL_SIZE_MAX) to catch affiliations rendered
    # in 7-9 pt by IEEEauthorblockA, which the old 10pt±0.6 filter missed.
    # Also exclude words that start past the page mid-point (right column)
    # to avoid false matches with body prose in correctly formatted papers.
    candidate_words = [
        w for w in words
        if w.get("top", 0) > last_author_top
        and w.get("top", 0) <= scan_end
        and AFFIL_SIZE_MIN <= _sz(w) <= AFFIL_SIZE_MAX
        and w.get("x0", 0) <= page_mid
    ]

    if not candidate_words:
        return True  # nothing suspicious

    # Build lines from those words
    candidate_words_sorted = sorted(candidate_words, key=lambda w: (round(w["top"]), w["x0"]))
    lines = _group_lines(candidate_words_sorted, tol=4.0)

    affil_lines_found = 0
    for line in lines:
        line_text = " ".join(w["text"] for w in line)
        if _line_is_affil(line_text):
            affil_lines_found += 1

    if affil_lines_found > 0:
        issues.append(_AFFIL_MSG)
        return False

    return True

# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------
def _extract_title(words: list, title_size: float, page_height: float) -> str:
    tw = sorted(
        [w for w in words
         if _matches(_sz(w), title_size, tol=1.5)
         and w.get("top", 0) < page_height * 0.45],
        key=lambda w: (round(w["top"]), w["x0"])
    )
    raw = " ".join(
        " ".join(w["text"] for w in ln)
        for ln in _group_lines(tw, 4)
    )
    raw = re.sub(r"-\s+", "", raw)
    raw = re.sub(r"\s+",  " ", raw).strip()
    raw = re.sub(r"^\d+\s+", "", raw).strip("*").strip()
    return _fix_encoding(raw)


def _is_name(text: str) -> bool:
    if not text or len(text) < 3:
        return False
    if RE_EMAIL.search(text) or RE_ORCID.search(text):
        return False
    words = text.split()
    if not (2 <= len(words) <= 8):
        return False
    if not (text[0].isupper() or text[0] in "ÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞß"):
        return False
    for word in words:
        base = re.sub(r"^[^a-zA-ZÀ-ÿ]+", "", word)
        if not base or base.lower() in _PREP:
            continue
        if not base[0].isupper():
            return False
        c = re.sub(r"[^a-záàâãéèêíïóôõúüçñ]", "", word.lower())
        if c in _AFFIL_KEYWORDS and c not in _PREP:
            return False
    return True


def _extract_authors(author_words: list) -> str:
    """Extract author names from the already-found author-line words."""
    if not author_words:
        return ""

    sorted_aw = sorted(author_words, key=lambda w: (round(w["top"]), w["x0"]))
    parts = []
    for line_words in _group_lines(sorted_aw, tol=4.0):
        line = RE_SUPER.sub(r"\1", " ".join(w["text"] for w in line_words)).strip()
        line = _fix_encoding(re.sub(r"\s+", " ", line).strip())
        if not line:
            continue
        # Split on "and" and commas, keep only name-like tokens
        for chunk in re.split(r"\band\b|,", line, flags=re.IGNORECASE):
            chunk = chunk.strip()
            if _is_name(chunk):
                parts.append(chunk)

    seen, unique = set(), []
    for name in parts:
        if name.lower() not in seen:
            seen.add(name.lower())
            unique.append(name)
    return ", ".join(unique)

# ---------------------------------------------------------------------------
# Main per-file processor
# ---------------------------------------------------------------------------
def process_pdf(pdf_path: Path) -> ValidationResult:
    result = ValidationResult(file=pdf_path.name)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                result.status = "INVALID"
                result.issues.append("PDF não possui páginas legíveis.")
                return result
            page      = pdf.pages[0]
            words     = page.extract_words(
                extra_attrs=["size", "fontname"],
                use_text_flow=False,
                keep_blank_chars=False,
            )
            page_text = page.extract_text() or ""
    except Exception as exc:
        result.status = "INVALID"
        result.issues.append(f"Não foi possível abrir ou analisar o PDF: {exc}")
        return result

    issues: list[str] = []

    # Structural checks
    _check_page_size(page, issues)
    _check_two_column(words, page.width, issues)
    _check_margins(words, page.width, issues)
    _check_abstract(page_text, issues)

    # Title check
    title_ok, title_size = _check_title_font(
        words, page.height, page.width, issues
    )

    # Author block checks (only if title was found)
    author_words: list = []
    last_author_top: float | None = None

    if title_size is not None:
        author_words, last_author_top = _find_author_line(
            words, title_size, page.height
        )
        if not author_words:
            issues.append(
                "Linha de autores não encontrada abaixo do título. "
                "Verifique o comando \\author{} no .tex."
            )
        else:
            _check_author_format(
                words, last_author_top, page.height, issues
            )

    # Final verdict
    if issues:
        result.status = "INVALID"
        result.issues = issues
        return result

    # Extraction
    result.status  = "VALID"
    result.title   = _extract_title(words, title_size, page.height)
    result.authors = _extract_authors(author_words)

    if not result.title:
        result.status = "INVALID"
        result.issues.append(
            "Passou nas verificações estruturais, mas o texto do título "
            "não pôde ser extraído."
        )
    if not result.authors:
        result.issues.append(
            "Título extraído; nomes dos autores não puderam ser isolados — "
            "verifique a formatação do \\author{} no .tex."
        )

    return result

# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------
def save_report(records: list, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RAS Format Validation"

    headers = ["ARQUIVO", "STATUS", "TÍTULO DO ARTIGO", "AUTORES", "MOTIVO DE NÃO-CONFORMIDADE"]
    thin   = Side(style="thin", color=COLORS["BORDER"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color=COLORS["HEADER_FT"])
        c.fill      = PatternFill("solid", start_color=COLORS["HEADER_BG"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    for ri, rec in enumerate(records, 2):
        valid = rec.status == "VALID"
        warn  = valid and bool(rec.issues)
        s_bg  = (COLORS["WARN_BG"]  if warn  else
                 COLORS["OK_BG"]    if valid else
                 COLORS["ERROR_BG"])
        s_ft  = (COLORS["WARN_FT"]  if warn  else
                 COLORS["OK_FT"]    if valid else
                 COLORS["ERROR_FT"])

        for ci, val in enumerate(
            [rec.file, rec.status, rec.title, rec.authors, rec.issues_text], 1
        ):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border    = border
            if ci == 2:
                c.fill = PatternFill("solid", start_color=s_bg)
                c.font = Font(name="Arial", size=9, bold=True, color=s_ft)
            elif ci == 5 and not valid:
                c.fill = PatternFill("solid", start_color=COLORS["ERROR_BG"])

    for ci, w in enumerate([30, 12, 62, 55, 95], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(output_path)

# ---------------------------------------------------------------------------
# Pipeline & CLI
# ---------------------------------------------------------------------------
def run_pipeline(pdf_folder: Path, output_xlsx: Path) -> None:
    if not pdf_folder.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {pdf_folder}")
    pdfs = sorted(pdf_folder.rglob("*.pdf"))
    if not pdfs:
        print(f"[AVISO] Nenhum arquivo PDF encontrado em: {pdf_folder}")
        return

    print(f"Checking {len(pdfs)} papers in '{pdf_folder}' for IEEE RAS template")

    records = []
    for pdf_path in pdfs:
        rec = process_pdf(pdf_path)
        records.append(rec)
        short  = pdf_path.name[:40]
        detail = (
            f"Título: {rec.title[:65]}{'...' if len(rec.title) > 65 else ''}"
            if rec.status == "VALID" and not rec.issues
            else f"[AVISO] {rec.issues[0]}" if rec.status == "VALID"
            else rec.issues[0][:90] if rec.issues else "Erro desconhecido"
        )

    valid = sum(1 for r in records if r.status == "VALID")
    print(f"{valid}/{len(records)} papers with RAS ieeeconf template.")
    save_report(records, output_xlsx)
    print(f"Report stored in: {output_xlsx}")


def main(folder) -> None:
    output = "./reports/ras_format_validation.xlsx"
    run_pipeline(Path(folder), Path(output))


if __name__ == "__main__":
    main()